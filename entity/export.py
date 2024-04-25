import os
import openpyxl
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font
from utils import get_today


def exportDataFrameEncaissement(df):
    wb = Workbook()
    ws = wb.active
    today = get_today()
    # Renommer la feuille de calcul
    ws.title = "ENCAISSEMENT"

    # Créer un dictionnaire pour stocker les données dans le format attendu
    formatted_data = {}
    for key, values in df.items():
        for value in values:
            intitule = value[0]
            amount = value[1]
            if intitule not in formatted_data:
                formatted_data[intitule] = {'day': 0, 'month': 0, 'year': 0}
            formatted_data[intitule][key] = amount

    # Écrire les en-têtes des colonnes avec mise en forme
    header_row = ['Intitulé', f'{today.strftime("%d/%m/%Y")}', 'Cumul Mois', 'Cumul Année']
    ws.append(header_row)
    for cell in ws[1]:
        cell.font = Font(bold=True)  # Rendre l'entête en gras

    # Écrire les données dans le fichier Excel avec mise en forme
    for intitule, amounts in formatted_data.items():
        ws.append([intitule, amounts['day'], amounts['month'], amounts['year']])

    # Ajouter une ligne TOTAL
    total_row = ['TOTAL', sum(amounts['day'] for amounts in formatted_data.values()),
                 sum(amounts['month'] for amounts in formatted_data.values()),
                 sum(amounts['year'] for amounts in formatted_data.values())]
    ws.append(total_row)

    # Ajouter des bordures aux cellules
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'),
                                 left=Side(style='thin'), right=Side(style='thin'))

    # Colorier les cellules de l'en-tête
    for cell in ws[1]:
        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    # Colorier la ligne TOTAL en bleu avec texte en blanc et gras
    for row in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.fill = PatternFill(start_color='0072BC', end_color='0072BC', fill_type='solid')
            cell.font = Font(color="FFFFFF", bold=True)

    # Espacez les colonnes selon les données
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Sauvegarder le fichier Excel
    storage_dir = f'storage/{get_today().strftime("%d-%m-%Y")}'
    if not os.path.exists(storage_dir):
        os.makedirs(storage_dir)

    # Sauvegarder le fichier Excel dans le dossier "storage"
    file_path = os.path.join(storage_dir, 'ENCAISSEMENT.xlsx')
    wb.save(file_path)

    # Retourner le chemin absolu du fichier
    return os.path.abspath(file_path)


def exportDataFrameDetails(df):
    # Créer un nouveau classeur Excel
    wb = openpyxl.Workbook()
    # Sélectionner la première feuille de calcul
    ws = wb.active
    ws.title = "VENTE DU JOUR DETAIL"

    # Ajouter l'en-tête de colonne
    for c_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col_name)
        cell.fill = openpyxl.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Ajouter les données du DataFrame à la feuille de calcul
    for r_idx, row in enumerate(df.iterrows(), start=2):
        for c_idx, value in enumerate(row[1], start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

    # Ajouter des bordures aux données
    for row in ws.iter_rows(min_row=1, max_row=len(df) + 1, min_col=1, max_col=len(df.columns)):
        for cell in row:
            cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                 right=openpyxl.styles.Side(style='thin'),
                                                 top=openpyxl.styles.Side(style='thin'),
                                                 bottom=openpyxl.styles.Side(style='thin'))

    # Ajouter une ligne de total à la fin pour la somme des colonnes D, E et F
    for c_idx, _ in enumerate(df.columns, start=1):
        if c_idx in [4, 5, 6]:  # Colonnes D, E et F (indices 3, 4 et 5)
            total = df.iloc[:, c_idx - 1].sum()
            cell = ws.cell(row=len(df) + 2, column=c_idx, value=total)

    # Ajouter le texte "TOTAUX" et formater la ligne
    totaux_row = len(df) + 2  # Ligne de sous-totaux
    ws.cell(row=totaux_row, column=1, value='TOTAUX').font = Font(bold=True, color='FFFFFF')
    for col in ws.iter_cols(min_row=totaux_row, max_row=totaux_row, min_col=1, max_col=len(df.columns)):
        for cell in col:
            cell.fill = openpyxl.styles.PatternFill(start_color="0072BC", end_color="0072BC", fill_type="solid")
            cell.font = Font(bold=True, color='FFFFFF')

    # Espacer les colonnes en fonction des données
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    # Créer le dossier de stockage s'il n'existe pas
    storage_dir = f'storage/{get_today().strftime("%d-%m-%Y")}'
    if not os.path.exists(storage_dir):
        os.makedirs(storage_dir)

    # Sauvegarder le fichier Excel dans le dossier "storage"
    file_path = os.path.join(storage_dir, 'DETAILS.xlsx')
    # Enregistrer le classeur Excel
    wb.save(file_path)
    return os.path.abspath(file_path)


def merge_excel_files(excel_1_path, excel_2_path, merged_file_path):
    # Charger les fichiers Excel existants
    wb1 = load_workbook(filename=excel_1_path)
    wb2 = load_workbook(filename=excel_2_path)

    # Créer un nouveau classeur Excel pour la fusion
    merged_wb = load_workbook()

    # Copier les feuilles de excel_1 dans le classeur fusionné
    for sheet_name in wb1.sheetnames:
        source = wb1[sheet_name]
        target = merged_wb.create_sheet(sheet_name)
        for row in source.iter_rows(min_row=1, max_row=source.max_row, min_col=1, max_col=source.max_column):
            for cell in row:
                target[cell.coordinate].value = cell.value

    # Copier les feuilles de excel_2 dans le classeur fusionné
    for sheet_name in wb2.sheetnames:
        source = wb2[sheet_name]
        target = merged_wb.create_sheet(sheet_name)
        for row in source.iter_rows(min_row=1, max_row=source.max_row, min_col=1, max_col=source.max_column):
            for cell in row:
                target[cell.coordinate].value = cell.value

    # Supprimer la feuille par défaut créée par load_workbook()
    merged_wb.remove(merged_wb['Sheet'])

    # Enregistrer le classeur fusionné
    merged_wb.save(filename=merged_file_path)

    return merged_file_path