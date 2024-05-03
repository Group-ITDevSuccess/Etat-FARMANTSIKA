import asyncio
import os
from copy import copy

import pyodbc
import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.workbook import Workbook

from entity.export import exportDataFrameEncaissement, exportDataFrameSimple, merge_excel_files
from utils import write_log, extract_values_in_json, get_today


def connexionSQlServer(server, base):
    conn = None
    value = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};DATABASE={base};UID=reader;PWD=m1234"
    print(value)
    try:
        conn = pyodbc.connect(value)
    except pyodbc.Error as e:
        # Si une erreur se produit lors de la connexion, ajoutez le serveur à la liste des bases sans accès
        print("===============================")
        print(f"Erreur de Connexion pour {server} à la base {base} ")
        print("===============================")
        write_log(f"Erreur de connexion : {str(e)}")
    except Exception as e:
        write_log(f"Erreur Exception : {str(e)}")
    return conn


def getDataLink(connexion, dates=None):
    print(dates)
    if connexion:
        try:
            finds = extract_values_in_json('SQL')
            if dates is None:
                dates = get_today()
            resumer = {}
            details = None
            pal_1 = None
            pal_2 = None
            for key, value in finds.items():
                if key not in ['details', 'palm_today', 'palm_all']:
                    with connexion.cursor() as cursor:
                        query = str(value).replace(
                            "{today}", dates.strftime('%m/%d/%Y')
                        ).replace("{year}", dates.strftime('%Y')
                                  ).replace('{month}', dates.strftime('%m'))
                        cursor.execute(query)
                        rows = cursor.fetchall()
                        tuples = []
                        if rows:
                            tuples = [tuple(row) for row in rows]
                        resumer[key] = tuples

                elif key in ['details', ]:
                    with connexion.cursor() as cursor:
                        query = str(value).replace("{today}", dates.strftime('%m/%d/%Y'))
                        cursor.execute(query)
                        rows = cursor.fetchall()
                        rows = [tuple(row) for row in rows]
                        if all(isinstance(row, tuple) for row in rows):
                            details = pd.DataFrame(rows,
                                                   columns=["N° PIECE", "REF", "DESIGNATION", "QTE", "PRIX UNITAIRE",
                                                            "MONTANT TTC",
                                                            "MODE"])
                        else:
                            print("Pas un tuple")
                else:
                    # print("On ast dans : "+key)
                    with connexion.cursor() as cursor:
                        try:
                            query = str(value).replace("{today}", dates.strftime('%m/%d/%Y'))
                            # print("-------------------------------")
                            # print(query)
                            # print("-------------------------------")
                            cursor.execute(query)
                            rows = cursor.fetchall()
                            rows = [tuple(row) for row in rows]
                            if all(isinstance(row, tuple) for row in rows):
                                if str(key).replace("palm_", '') == 'today':
                                    pal_1 = pd.DataFrame(rows,
                                                         columns=["REF", "DESIGNATION", "QUANTITE", "MONTANT TTC"])
                                    # print(f"PALM 1 : {pal_1}")
                                elif str(key).replace("palm_", '') == 'all':
                                    pal_2 = pd.DataFrame(rows,
                                                         columns=["ARTICLE", "DESIGNATION", "QUANTITE", "MONTANT TTC"])
                                    # print(f"PALM 2 : {pal_2}")
                        except Exception as e:
                            print(f"Erreur : {key}")
                            write_log(str(e))
                            pass
            excel_1 = exportDataFrameEncaissement(resumer, dates)
            excel_2 = exportDataFrameSimple(details, f"VENTE DU JOUR DETAIL {dates.strftime("%d-%m-%Y")}",
                                            [4, 5, 6])
            excel_3 = exportDataFrameSimple(pal_1, f"PALMARES PRODUIT {dates.strftime("%d-%m-%Y")}", [3, 4])
            excel_4 = exportDataFrameSimple(pal_2, f"PALMARES PRODUIT DEPUIS", [3, 4])

            storage_dir = f'storage/{dates.strftime("%d-%m-%Y")}'
            if not os.path.exists(storage_dir):
                os.makedirs(storage_dir)
            wb1 = load_workbook(excel_1)
            wb2 = None
            wb3 = None
            wb4 = None
            if excel_2 is not None:
                wb2 = load_workbook(excel_2)
            if excel_3 is not None:
                wb3 = load_workbook(excel_3)
            if excel_4 is not None:
                wb4 = load_workbook(excel_4)

            def copie_excel(excel_begin, excel_inject):
                for sheet in excel_inject.sheetnames:
                    ws = excel_inject[sheet]
                    if sheet not in excel_begin.sheetnames:
                        excel_begin.create_sheet(title=sheet)

                    # Copier les styles de l'Excel_2 vers l'Excel_1
                    for row in ws.iter_rows():
                        for cell in row:
                            dest_cell = excel_begin[sheet][cell.coordinate]
                            dest_cell.font = copy(cell.font)
                            dest_cell.fill = copy(cell.fill)
                            dest_cell.border = copy(cell.border)
                            dest_cell.alignment = copy(cell.alignment)
                            dest_cell.number_format = copy(cell.number_format)
                            dest_cell.protection = copy(cell.protection)

                    # Copier les données de l'Excel_2 vers l'Excel_1 dans les mêmes cellules
                    for row in ws.iter_rows():
                        for cell in row:
                            dest_cell = excel_begin[sheet][cell.coordinate]
                            dest_cell.value = cell.value

                # Espacer les colonnes en fonction du contenu
                for sheet in excel_begin.sheetnames:
                    ws = excel_begin[sheet]
                    for column_cells in ws.columns:
                        max_length = 0
                        for cell in column_cells:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
                            except:
                                pass
                        adjusted_width = (max_length + 2) * 1.2
                        ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

                # Supprimer la première feuille vide créée automatiquement dans excel_begin
                if 'Sheet' in excel_begin.sheetnames:
                    excel_begin.remove(excel_begin['Sheet'])
                return excel_begin

            if wb2:
                wb1 = copie_excel(wb1, wb2)
            if wb3:
                wb1 = copie_excel(wb1, wb3)
            if wb4:
                wb1 = copie_excel(wb1, wb4)

            storage_out = f'storage/{dates.strftime("%d-%m-%Y")}/output'
            if not os.path.exists(storage_out):
                os.makedirs(storage_out)
            merged_file_path = os.path.join(storage_out,
                                            f'ETAT THE MEAT SHOP {dates.strftime("%d-%m-%Y")}.xlsx')
            wb1.save(merged_file_path)

            print(f"Fichier fusionné enregistré à : {merged_file_path}")

            return merged_file_path
        except Exception as e:
            write_log(f"Erreur Exception : {str(e)}")
        finally:
            connexion.close()

        return None
